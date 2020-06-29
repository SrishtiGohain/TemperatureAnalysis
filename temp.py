import openpyxl 
import tkinter as tk
from tkinter import ttk

from openpyxl import load_workbook

destination='C:/Users/Srishti/Desktop/city_temperature.xlsx'
workbook=load_workbook(destination)
worksheet=workbook.active

def print_temp(city,month,day,year):
    for r in range(2,1048576):
        c=worksheet.cell(row=r,column=4)
        m=worksheet.cell(row=r,column=5)
        d=worksheet.cell(row=r,column=6)
        y=worksheet.cell(row=r,column=7)
        if c.value==city and m.value==month and d.value==day and y.value==year:
            print(worksheet.cell(row=r+1,column=8).value)
            break

print_temp("Algiers",1,1,1995)
root=tk.Tk()
root.title("Global City Temperatures")
heading=ttk.Label(root,text="Next Day Temperature")
city_name=ttk.Label(root, text="Your City Name")
month_val=ttk.Label(root, text="Month")
day_val=ttk.Label(root, text="Day")
year_val=ttk.Label(root, text="Year")

heading.grid()
city_name.grid(row=1, column=0)
month_val.grid(row=2,column=0)
day_val.grid(row=3, column=0)
year_val.grid(row=4,column=0)

root.mainloop()


